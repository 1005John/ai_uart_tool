#!/usr/bin/env python3
"""Excel用例加载器 - 读取 test_at 框架的 xlsx 测试配置文件"""
import os
import re
import ast
import json
from typing import Optional

# 优先使用 project venv 的 openpyxl
try:
    import openpyxl
except ImportError:
    import sys
    # fallback to user install
    sys.path.insert(0, os.path.expanduser("~/.local/lib/python3.11/site-packages"))
    import openpyxl

from models.schemas import TestCase


# ── 模块 → Excel 文件映射 ──
MODULE_MAP = {
    "mqtt": {
        "file": "mqtt_test_configs.xlsx",
        "commands": "command_mqtt.py",
        "description": "MQTT协议测试",
    },
    "mqttcfg": {
        "file": "mqttcfg_test_configs.xlsx",
        "commands": "command_mqtt.py",
        "description": "MQTT配置测试",
    },
    "tcp": {
        "file": "tcp_test_configs.xlsx",
        "commands": "command_tcp.py",
        "description": "TCP协议测试",
    },
    "http": {
        "file": "http_test_configs.xlsx",
        "commands": "command_http.py",
        "description": "HTTP协议测试",
    },
    "sms": {
        "file": "sms_test_configs.xlsx",
        "commands": "command_sms.py",
        "description": "短信功能测试",
    },
    "gnss": {
        "file": "gnss_test_configs.xlsx",
        "commands": "command_gnss.py",
        "description": "GNSS定位测试",
    },
    "fota": {
        "file": "fotacfg_test_configs.xlsx",
        "commands": "command_fota.py",
        "description": "FOTA升级测试",
    },
    "ssl": {
        "file": "mssl_test_configs.xlsx",
        "commands": "command_ssl.py",
        "description": "SSL/TLS测试",
    },
    "dns": {
        "file": "dns_test_configs.xlsx",
        "commands": "command_dns.py",
        "description": "DNS解析测试",
    },
    "sim": {
        "file": "sms_test_configs.xlsx",  # Note: check actual file
        "commands": "command_sim.py",
        "description": "SIM卡测试",
    },
    "dialup": {
        "file": "dialup_test_configs.xlsx",
        "commands": "command_dialup.py",
        "description": "拨号测试",
    },
    "wifi": {
        "file": "wifi_test_configs.xlsx",
        "commands": "command_wifi.py",
        "description": "WiFi测试",
    },
    "gpio": {
        "file": "mgpio_test_configs.xlsx",
        "commands": "command_gpio.py",
        "description": "GPIO测试",
    },
    "ntp": {
        "file": "ntp_test_configs.xlsx",
        "commands": "command_ntp.py",
        "description": "NTP时间同步",
    },
    "ftp": {
        "file": "mftp_test_configs.xlsx",
        "commands": "command_ftp.py",
        "description": "FTP文件传输",
    },
    "networkservice": {
        "file": "networkservice_test_configs.xlsx",
        "commands": "command_networkservice.py",
        "description": "网络服务测试",
    },
    "packet_domain": {
        "file": "packet_domain_test_configs.xlsx",
        "commands": "command_packet_domain.py",
        "description": "PDP上下文测试",
    },
    "dmp": {
        "file": "dmp_test_configs.xlsx",
        "commands": "command_dm.py",
        "description": "设备管理测试",
    },
    "adc": {
        "file": "adc_test_configs.xlsx",
        "commands": "command_adc.py",
        "description": "ADC测试",
    },
    "pm": {
        "file": "pmcfg_test_configs.xlsx",
        "commands": "command_pm.py",
        "description": "电源管理测试",
    },
    "me_control": {
        "file": "me_control_command_test_configs.xlsx",
        "commands": "command_me_control.py",
        "description": "ME控制测试",
    },
    "extended": {
        "file": "extended_command_test_configs.xlsx",
        "commands": "command_extended.py",
        "description": "扩展AT指令测试",
    },
    "mled": {
        "file": "mled_test_configs.xlsx",
        "commands": "command_mled.py",
        "description": "LED灯测试",
    },
    "mping": {
        "file": "mping_test_configs.xlsx",
        "commands": "command_mping.py",
        "description": "网络Ping测试",
    },
}


class ExcelTestLoader:
    """
    Excel测试用例加载器

    读取 test_at 框架的 xlsx 测试配置文件，
    按模块/Sheet/行解析为 TestCase 对象。
    """

    def __init__(self, data_dir: str = None):
        """
        Args:
            data_dir: 包含 xlsx 文件的目录，默认自动查找
        """
        self.data_dir = data_dir or self._find_data_dir()
        self._wb_cache = {}

    def _find_data_dir(self) -> str:
        """自动查找 test_data_configs 目录"""
        candidates = [
            # 项目内链接
            os.path.expanduser("~/ai_uart_tool/test_data"),
            # 原始 test_at 仓库位置
            os.path.expanduser("~/test_at/test_data_configs"),
            os.path.expanduser("~/code/test_at/test_data_configs"),
            "/tmp/test_at_repo/test_at/test_data_configs",
            # 当前目录
            os.path.join(os.getcwd(), "test_data_configs"),
            os.path.join(os.path.dirname(os.path.dirname(__file__)), "test_data"),
        ]
        for path in candidates:
            if os.path.isdir(path):
                print(f"[ExcelLoader] 使用数据目录: {path}")
                return path
        print(f"[ExcelLoader] ⚠️ 未找到测试数据目录，使用: {candidates[0]}")
        os.makedirs(candidates[0], exist_ok=True)
        return candidates[0]

    def list_modules(self) -> list[dict]:
        """列出所有可用模块"""
        result = []
        for key, info in MODULE_MAP.items():
            xlsx_path = os.path.join(self.data_dir, info['file'])
            if os.path.exists(xlsx_path):
                result.append({
                    "key": key,
                    "file": info['file'],
                    "description": info['description'],
                    "path": xlsx_path,
                })
        return result

    def get_module_info(self, module: str) -> dict:
        """获取模块信息"""
        if module not in MODULE_MAP:
            raise ValueError(f"未知模块: {module}，可用: {list(MODULE_MAP.keys())}")
        info = MODULE_MAP[module].copy()
        info['path'] = os.path.join(self.data_dir, info['file'])
        return info

    def load_sheets(self, module: str) -> list[str]:
        """加载指定模块的所有Sheet名称"""
        info = self.get_module_info(module)
        wb = self._get_workbook(info['path'])
        return wb.sheetnames

    def load_cases(self, module: str, sheet_name: str = None,
                   model: str = None, case_level: list = None) -> list[TestCase]:
        """
        加载测试用例

        Args:
            module: 模块名 (mqtt, tcp, http...)
            sheet_name: Sheet名，None则加载所有Sheet
            model: 模组型号过滤 (ML307C-DC...)
            case_level: 用例级别过滤 (['P0'], ['P0','P1']...)

        Returns: TestCase列表
        """
        info = self.get_module_info(module)
        wb = self._get_workbook(info['path'])

        all_cases = []
        sheets = [sheet_name] if sheet_name else wb.sheetnames

        for sn in sheets:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]

            # 读取表头
            headers = []
            if ws.max_row is None or ws.max_row < 1 or ws.max_column is None or ws.max_column < 1:
                continue
            first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            if not first_row or not first_row[0]:
                continue
            for cell_val in first_row[0]:
                if cell_val is not None:
                    cell_str = str(cell_val).strip()
                    if cell_str:
                        headers.append(self._normalize_header(cell_str))
            if not headers:
                continue

            # 逐行读取
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if all(v is None for v in row):
                    continue

                case = self._row_to_case(info['path'], sn, row_idx, headers, row)
                if case is None:
                    continue

                # 模组过滤
                if model and not self._model_matches(case.applicable_models, model):
                    continue

                # 级别过滤
                if case_level and case.case_level not in case_level:
                    continue

                all_cases.append(case)

        return all_cases

    def _get_workbook(self, path: str):
        """获取缓存的Workbook"""
        if path not in self._wb_cache:
            if not os.path.exists(path):
                raise FileNotFoundError(f"Excel文件不存在: {path}")
            self._wb_cache[path] = openpyxl.load_workbook(path, read_only=True, data_only=True)
        return self._wb_cache[path]

    def _normalize_header(self, header: str) -> str:
        """标准化表头名称"""
        mapping = {
            'id': 'id',
            'test_group': 'test_group',
            'name': 'case_name',
            'case_name': 'case_name',
            'projects': 'applicable_models',
            'case_level': 'case_level',
            'expected_results': 'expected_results',
        }
        return mapping.get(header, header)

    def _row_to_case(self, file_path: str, sheet_name: str, row_idx: int,
                     headers: list, row: tuple) -> Optional[TestCase]:
        """Excel行 → TestCase"""
        row_dict = {}
        for i, h in enumerate(headers):
            if i < len(row):
                row_dict[h] = row[i]

        # 必填字段检查
        case_name = self._safe_str(row_dict.get('case_name', ''))
        if not case_name:
            return None

        # 解析适用模组
        projects_raw = row_dict.get('applicable_models', '["ALL"]')
        applicable_models = self._parse_model_list(projects_raw)

        # 解析预期结果
        expected_raw = row_dict.get('expected_results', '[]')
        expected = self._parse_list(expected_raw)

        # 级别
        case_level = self._safe_str(row_dict.get('case_level', 'P0'))
        if not case_level:
            case_level = 'P0'

        # 提取 AT 指令相关参数（第一列通常是ID，跳过）
        at_params = {}
        for h in headers:
            if h not in ('id', 'test_group', 'case_name', 'applicable_models',
                         'case_level', 'expected_results'):
                at_params[h] = row_dict.get(h)

        return TestCase(
            excel_file=file_path,
            sheet_name=sheet_name,
            row_id=row_idx,
            test_group=self._safe_str(row_dict.get('test_group', '')),
            case_name=case_name,
            case_level=case_level.upper(),
            applicable_models=applicable_models,
            params=at_params,
            expected_results=expected,
        )

    def _safe_str(self, val) -> str:
        if val is None:
            return ""
        return str(val).strip()

    def _parse_model_list(self, val) -> list:
        """解析 projects 字段"""
        if val is None:
            return ["ALL"]
        s = str(val).strip()
        if not s:
            return ["ALL"]
        try:
            return ast.literal_eval(s)
        except:
            return [s]

    def _parse_list(self, val) -> list:
        """解析列表字段"""
        if val is None:
            return []
        s = str(val).strip()
        if not s:
            return []
        try:
            return ast.literal_eval(s)
        except:
            return [s]

    def _model_matches(self, model_list: list, target_model: str) -> bool:
        """判断模组型号是否匹配"""
        for entry in model_list:
            if entry == "ALL":
                return True
            if entry.startswith("ALL|"):
                excluded = entry.split("|")[1].split("/")
                if target_model in excluded:
                    continue
                return True
            if target_model == entry:
                return True
        return False

    def cases_to_text(self, cases: list[TestCase]) -> str:
        """用例列表 → 可读文本"""
        if not cases:
            return "无匹配用例"
        lines = [f"共 {len(cases)} 条用例:"]
        for c in cases:
            checked = "☑" if c.selected else "☐"
            level = c.case_level
            lines.append(f"  {checked} [{level}] {c.case_name}")
            for k, v in c.params.items():
                if v is not None and str(v).strip():
                    lines.append(f"       {k}={v}")
            if c.expected_results:
                lines.append(f"       预期: {c.expected_results}")
        return "\n".join(lines)


# ── 便捷工具函数 ──

def _safe_eval_str(s: str):
    """安全解析字符串表达式，支持 "a"*128 等"""
    s = s.strip()
    if s.startswith('"') and s.count('"') == 2 and '*' in s:
        quote_char = s[0]
        second_quote = s.find(quote_char, 1)
        if second_quote > 0:
            asterisk = s.find('*', second_quote)
            if asterisk > 0:
                content = s[1:second_quote]
                num = s[asterisk + 1:].strip()
                if num.isdigit():
                    return content * int(num)
    return s


def evaluate_param(value, model: str = None) -> str:
    """评估参数表达式，返回实际值"""
    if value is None:
        return ""
    s = str(value)
    result = _safe_eval_str(s)
    return result


if __name__ == "__main__":
    loader = ExcelTestLoader()
    print("可用模块:", [m['key'] for m in loader.list_modules()])

    if loader.list_modules():
        mod = loader.list_modules()[0]['key']
        print(f"\n加载模块: {mod}")
        sheets = loader.load_sheets(mod)
        print(f"  Sheets: {sheets}")
        cases = loader.load_cases(mod)
        print(f"\n{loader.cases_to_text(cases[:10])}")
        if len(cases) > 10:
            print(f"  ... 还有 {len(cases)-10} 条")
